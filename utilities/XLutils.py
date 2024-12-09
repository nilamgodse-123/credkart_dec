
import openpyxl

def RowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row

def ReadData(file,sheetname,row_no,colu_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=row_no,column=colu_no).value

def WriteData(file,sheetname,row_no,colu_no,data):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row=row_no,column=colu_no).value=data
    workbook.save(file)